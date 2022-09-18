from openpyxl import load_workbook

import os 
import cx_Oracle
from pathlib import Path
import time

#filepath = "./limit.xlsx"
#sfile = r"C:\temp\SIP_PARA\SAMPLE\Younger_limit.xlsx"
source_dir = "C:/temp/SIP_PARA/"

limit_dir = source_dir + "LIMIT/"

msg = 'Limit file Insert start'
print(msg)
                    
# send_mail('mail test' )
p = Path(source_dir)  # 로컬 다운받은 폴더
p.mkdir(exist_ok=True)
p = Path(limit_dir)  # 로컬 다운받은 폴더
p.mkdir(exist_ok=True)
 
for (path, dir, files) in os.walk(limit_dir):
    for filename in files: 
        
        ext = os.path.splitext(filename)[-1]
        path = os.path.abspath(path)
        full_filename = os.path.join(path, filename)
        
        if (ext == '.xlsx') and filename.upper().find('LIMIT') > -1 :
        
            wb = load_workbook(full_filename)
            ws = wb.active

            print(wb.sheetnames) 
            project = ws['A1'].value
            project = project.upper()

            dic_limit = {}
            dic_limit.clear()

            data = []

            for i in range(1, ws.max_row + 1) :    	
                if i == 2 :
                    if (ws['B2'].value.upper() != 'SERIALNUMBER' or ws['C2'].value.upper() != 'UPPER LIMIT' or  ws['D2'].value.upper() != 'LOWER LIMIT' ) :
                        print("header 가 일치하지 않습니다.")
                for j in range(1, ws.max_column + 1) :
                    # get particular cell value
                    print(ws.cell(row=i, column=j).value, end= '~')
                    
                    add_data = (ws.cell(row=i, column=j).value)
                    
                    if j == 2 :
                        spara_name = str(ws.cell(row=i, column=j).value)
                    if j == 3 :
                        upper_limit = str(ws.cell(row=i, column=j).value)
                    if j == 4 :
                        lower_limit = str(ws.cell(row=i, column=j).value)
                print('\n') 
            
                if (i > 2) and  (upper_limit != "NA" or lower_limit != "NA") :
                    add_data = (project , 'NA' , spara_name, upper_limit, lower_limit , filename )
                    data.append(add_data)
                    print(add_data)
            ##########################################################################
            sip_conn = cx_Oracle.connect('sipdb')
            sip_cur = sip_conn.cursor()
            if len(data) > 0:    
                try :
                    #----------------------------------------------
                    # 동일 project 명으로 있으면 삭제후 insert    
                    sql = "SELECT PROJECT FROM SIP_PARA_LIMIT "
                    sql = sql + "WHERE PROJECT = '" + project + "' "
                    sip_cur.execute(sql)
                    rows = sip_cur.fetchall()
                    if  sip_cur.rowcount > 0 :
                        sql = "DELETE FROM SIP_PARA_LIMIT "
                        sql = sql + "WHERE PROJECT = '" + project + "' "
                        sip_cur.execute(sql)
                        sip_conn.commit()
                    #-----------------------------------------
                    sql = "INSERT INTO SIP_PARA_LIMIT (PROJECT, GUBUN, PARA_NAME, UPPER_LIMIT, LOWER_LIMIT, FILENAME, TRANS_TIME ) "
                    #                                   1     ,   2  ,  3       ,  4         ,      5     ,      6  ,  7  
                    sql = sql + "VALUES ( :1, :2, :3, :4, :5, :6, to_char(sysdate,'yyyymmddhh24miss') ) "

                    sip_cur.executemany(sql,data)            
                    sip_conn.commit()
                        
                    msg = str(i) + ' ] s5. SIP_PARA_LIMIT Insert complete !! ' + '\n' + " " + filename
                    print(msg)
                    time.sleep(3)
                    
                except Exception as e:  # 예외가 발생했을 때 실행됨                     
                    msg = 'Limit file Insert Error !!' + '\n' + e + '\n' + filename
                    print(msg)
                    time.sleep(3)
                
                ##################   
